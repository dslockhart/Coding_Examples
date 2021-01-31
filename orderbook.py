# Import relevant libraries
import pandas.io.sql
import pandas as pd
import numpy as np
import pyodbc
import re
import openpyxl
from datetime import datetime

# Required formatting for Pandas
desired_width = 320
pd.set_option('display.width', desired_width)
np.set_printoptions(linewidth=desired_width)
pd.set_option('display.max_columns', 10)

# Define variables and empty lists
customerdict = {'&CDADDER': ['SAFRAN', 'SAFRANOB.xlsx'], 'AIMAVIAT': ['AIM', 'AIMOB.xlsx']}  #default customer variables
col_no_dict = {'AIMAVIAT': [12, 16], '&CDADDER': [10, 13]}  # orderbook date column and output date column
date_dict = {'AIMAVIAT': '%d/%m/%Y %H:%M:%S', '&CDADDER': '%m/%d/%y'}  # default customer date formats - note can vary

# Parsing user input for customer account no
while True:
    cus_no = input('What is the customer account no?:\n')
    cus_no = cus_no.upper()
    if cus_no not in customerdict.keys():
        print('Customer not found')
        print('*'*100)
    else:
        print('Completing orderbook for {}...'.format(cus_no))
        break
# Continuation of variable declarations
cus_name = customerdict[cus_no][0]  # Used to save output file name
ob_name = customerdict[cus_no][1]  # Enter the Excel document name w/ .xlsx
cus_col_po = 1  # Which column is PO number
cus_col_ln = 2  # Which column is PO line number
cus_col_date = col_no_dict[cus_no][0]  # Which column has line delivery date
cus_output_date = col_no_dict[cus_no][1]  # Which column does supplier input response
po_no = []
po_line = {}
nan = [] # blank list for orderbook lines that are not in javelin

# SQL connection and queries
conx_string = 'DRIVER={SQL Server};SERVER=NMC-JAV; DATABASE=Javelin19r1_Live; '
'TRUSTED CONNECTION=yes'
query = "SELECT Del_Note_Detail.Del_Note_No, Del_Note_Detail.SO_No, " \
        "CONCAT(Del_Note_Detail.SO_No, '/', Del_Note_Detail.SO_Line_No) AS Concat," \
        "Del_Note_Detail.SO_Line_No, Del_Note_Footer.Text FROM " \
        "despatch.Del_Note_Detail JOIN despatch.Del_Note_Footer ON Del_Note_Detail.Del_Note_No = Del_Note_Footer.Del_Note_No;" \
 \
query2 = "SELECT SO_Items.Customer_No, SO_Items.SO_No," \
         "CONCAT(SO_Header.Customer_PO_No, '/', SO_Items.SO_Line_No) AS Concat," \
         "CONCAT(SO_Header.SO_No, '/', SO_Items.SO_Line_No) AS C1," \
         " SO_Items.Qty, SO_Items.Qty_Del, SO_Items.Promise_Date" \
         " FROM sales.SO_Items LEFT JOIN sales.SO_Header ON SO_Items.SO_No = SO_Header.SO_No " \
         "WHERE SO_Items.Customer_No = '{}';".format(cus_no)

with pyodbc.connect(conx_string) as conx:
    cursor = conx.cursor()
    cursor.execute(query)

# Defining Pandas dataframes
df = pandas.io.sql.read_sql(query, conx)
df2 = pandas.io.sql.read_sql(query2, conx)
df.set_index('Concat')
df2.set_index('Concat')

# Regex search for tracking numbers
df = df[df['Text'].str.contains(re.compile(r'\d{7,}'))]
df = df.drop_duplicates(subset="Concat", keep='first')



# Combining dataframes from SQL queries
df_comb = df2.merge(df, left_on="C1", right_on="Concat", how='left')
df_comb['clean_date'] = df_comb['Promise_Date'].apply(lambda x: x.strftime('%Y-%m-%d'))

# Open customer orderbook from Excel
wb = openpyxl.load_workbook(ob_name)
sheet = wb["Sheet1"]

for i in range(2, sheet.max_row+1):
    po = sheet.cell(row=i, column=cus_col_po).value.strip()
    po_no.append(po)
    line = int(sheet.cell(row=i, column=cus_col_ln).value)
    a = (po + "/" + str(line))
    date = sheet.cell(row=i, column=cus_col_date).value
    if isinstance(date, str):
        date_obj = datetime.strptime(date, '{}'.format(date_dict[cus_no])).date()
    else:
        date = datetime.strftime(date, '%d/%m/%y')
        date_obj = datetime.strptime(date, '%m/%d/%y')     # Note Safran has varied this format
    po_line[a] = []
    po_line[a].append(date_obj)

# Creates a dataframe from customer orderbook
dfob = pd.DataFrame(list(po_line.items()), columns=['Concat', 'OB_date'])

# Cleans orderbook dates into string
dfob['ob_clean'] = dfob['OB_date'].apply(lambda x: x[0].strftime('%Y-%m-%d'))

# Complete dataframe of javelin and orderbook
data = df_comb.merge(dfob, left_on="Concat_x", right_on='Concat', how='right')

# Tidy up columns
data = data.drop(columns=['SO_No_x', 'C1', 'Del_Note_No', 'SO_No_y',
                          'Concat_y', 'SO_Line_No', 'Concat', 'OB_date'])

# Further date cleaning
data['ob_clean'] = data['ob_clean'].astype(str)
data['Comparison'] = (data['clean_date'] == (data['ob_clean']))

# variable used for checking if date in past
a = datetime.now()


# loops through orderbook values and autocompletes based on date criteria
for i, j in data.iterrows():
    if (data['Comparison'][i]) and data['Promise_Date'][i] < a: # if due in past
        try:
            sheet.cell(row=i+2, column=cus_output_date).value = "Shipped: {}; {} ({} Remaining)" \
                .format(int(data['Qty_Del'][i]), data['Text'][i], (int(data['Qty'][i]))-int(data['Qty_Del'][i]))
        except:
            continue
    elif not (data['Comparison'][i]) and data['Promise_Date'][i] < a and str(data['Text']) != 'nan':
        sheet.cell(row=i+2, column=cus_output_date).value = "SHIPPED LATE: {}".format(str(data['Text'][i]))

    elif (data['Comparison'][i]) and data['Promise_Date'][i] > a:
        try:
            sheet.cell(row=i+2, column=cus_output_date).value = "Expected on time: {}".format(data['clean_date'][i])
        except:
            continue

    elif str(data['clean_date'][i]) != 'nan':
        sheet.cell(row=i+2, column=cus_output_date).value = "Expected: {}".format(data['clean_date'][i])

    else:
        sheet.cell(row=i+2, column=cus_output_date).value = "Line error"
wb.save('{}_Completed_orderbook.xlsx'.format(cus_name))